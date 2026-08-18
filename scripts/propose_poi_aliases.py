"""Offline: propose candidate POI alias groups from the CAISO workbook, for human review.

This is the *only* place fuzzy matching is allowed. It reads the active sheet's free-text
station strings, normalizes them, and clusters the normalized keys by string similarity
(rapidfuzz) so a reviewer sees likely-same substations grouped together. The output is a
proposal, not an answer: a human edits it down to the reviewed
``src/interconnection_agent/poi/aliases.csv`` that the ETL applies by exact match.

It is a developer-facing script, deliberately outside the ``interconnection_agent``
package and never imported by it — a probabilistic join must never run beneath a verified
claim (``test_poi_offline_only`` enforces the boundary). Run it directly:

    uv run python scripts/propose_poi_aliases.py data/publicqueuereport.xlsx > proposal.csv

Each output row is ``group_id,station,normalized_key,n,mw,proposed_canonical``. Rows
sharing a ``group_id`` are the tool's guess at one POI; ``proposed_canonical`` seeds the
canonical name (the highest-MW member of the group). The reviewer confirms, splits, or
merges groups and keeps only ``station,canonical_poi`` for the checked-in table.
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz

# Import the runtime normalizer so the proposal is keyed exactly as the ETL will key it.
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
from interconnection_agent.poi import normalize_station  # noqa: E402

ACTIVE_SHEET = "Grid GenerationQueue"
HEADER_ROW = 4
FIRST_DATA_ROW = 5
STATION_HEADER = "Station or Transmission Line"
STATUS_HEADER = "Application Status"
NET_MW_HEADER = "Net MWs to Grid"

# Two normalized keys join a group when their token-sort similarity clears this. Tuned for
# review generosity: better to over-group and let a human split than to hide a real alias.
SIMILARITY_THRESHOLD = 88.0


@dataclass
class Station:
    raw: str
    key: str
    n: int = 0
    mw: float = 0.0


@dataclass
class Group:
    members: list[Station] = field(default_factory=list)

    @property
    def mw(self) -> float:
        return sum(s.mw for s in self.members)

    def canonical(self) -> str:
        """Seed the canonical name from the highest-MW member's raw spelling."""
        return max(self.members, key=lambda s: s.mw).raw


def _norm_header(value: object) -> str:
    return " ".join(str(value).split()) if value is not None else ""


def read_active_stations(workbook_path: Path) -> list[Station]:
    """Collect distinct active-sheet station strings with their project count and MW."""
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[ACTIVE_SHEET]
        header = next(sheet.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
        col = {_norm_header(v): i for i, v in enumerate(header) if v is not None}
        by_key: dict[str, Station] = {}
        for row in sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
            if all(v is None for v in row):
                continue
            status = row[col[STATUS_HEADER]]
            if status is None or str(status).strip().upper() != "ACTIVE":
                continue
            raw = row[col[STATION_HEADER]]
            if raw is None or not str(raw).strip():
                continue
            mw = row[col[NET_MW_HEADER]]
            key = normalize_station(raw)
            station = by_key.setdefault(key, Station(raw=str(raw).strip(), key=key))
            station.n += 1
            station.mw += float(mw) if mw is not None else 0.0
        return sorted(by_key.values(), key=lambda s: -s.mw)
    finally:
        workbook.close()


def cluster(stations: list[Station], threshold: float = SIMILARITY_THRESHOLD) -> list[Group]:
    """Greedily group stations whose normalized keys are similar enough to review together."""
    groups: list[Group] = []
    for station in stations:
        for group in groups:
            if any(fuzz.token_sort_ratio(station.key, m.key) >= threshold for m in group.members):
                group.members.append(station)
                break
        else:
            groups.append(Group(members=[station]))
    return sorted(groups, key=lambda g: -g.mw)


def write_proposal(groups: list[Group], out: object) -> None:
    writer = csv.writer(out)
    writer.writerow(["group_id", "station", "normalized_key", "n", "mw", "proposed_canonical"])
    for group_id, group in enumerate(groups, start=1):
        canonical = group.canonical()
        for member in sorted(group.members, key=lambda s: -s.mw):
            writer.writerow(
                [group_id, member.raw, member.key, member.n, round(member.mw, 1), canonical]
            )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Propose POI alias groups for human review.")
    parser.add_argument("workbook", type=Path, help="path to publicqueuereport.xlsx")
    parser.add_argument(
        "--threshold", type=float, default=SIMILARITY_THRESHOLD, help="similarity cutoff (0-100)"
    )
    args = parser.parse_args(argv)
    stations = read_active_stations(args.workbook)
    groups = cluster(stations, args.threshold)
    write_proposal(groups, sys.stdout)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
